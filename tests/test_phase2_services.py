import unittest
from datetime import date
from decimal import Decimal
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import load_workbook

from app.models.bank_template import BankTemplate
from app.models.enums import DuplicateStatus, ValidationStatus, VoucherType
from app.models.mapping import ColumnMapping
from app.models.transaction import Transaction
from app.services.bank_template_service import BankTemplateService
from app.services.duplicate_detector import DuplicateDetector
from app.services.excel_exporter import ExcelExporter
from app.services.transaction_normalizer import TransactionNormalizer
from app.services.transaction_parser import TransactionParser
from app.services.validation_service import ValidationService


def transaction(description: str = "UPI DEMO TRADERS") -> Transaction:
    return Transaction(
        transaction_date=date(2026, 7, 1),
        description=description,
        reference_number="UTR001",
        debit_amount=Decimal("1250.50"),
        balance=Decimal("8749.50"),
        voucher_type=VoucherType.PAYMENT,
    )


class PhaseTwoServiceTests(unittest.TestCase):
    def test_multiline_narration_is_merged(self) -> None:
        rows = [
            ["Date", "Narration", "Debit", "Credit", "Balance"],
            ["01/07/2026", "UPI PAYMENT TO", "1,250.50", "", "8,749.50"],
            ["", "DEMO TRADERS UTR001", "", "", ""],
            ["02/07/2026", "INTEREST", "", "12.75", "8,762.25"],
        ]
        mapping = ColumnMapping(transaction_date=0, description=1, debit=2, credit=3, balance=4)
        template = BankTemplate("Test Bank", "test", multiline_enabled=True)
        items, warnings = TransactionParser().parse(rows, mapping, template)
        self.assertFalse(warnings)
        self.assertEqual(len(items), 2)
        self.assertEqual(items[0].description, "UPI PAYMENT TO DEMO TRADERS UTR001")
        self.assertEqual(items[0].narration, items[0].description)

    def test_amount_normalization_supports_suffix_and_negative_debits(self) -> None:
        self.assertEqual(
            TransactionNormalizer.normalize(amount_raw="500.00CR"),
            (Decimal("0.00"), Decimal("500.00")),
        )
        self.assertEqual(
            TransactionNormalizer.normalize(debit_raw="-125.00"),
            (Decimal("125.00"), Decimal("0.00")),
        )

    def test_duplicate_detection_and_user_unique_override(self) -> None:
        first = transaction()
        second = transaction()
        detector = DuplicateDetector()
        matches = detector.detect([first, second], 75)
        self.assertEqual(len(matches), 1)
        self.assertEqual(second.duplicate_status, DuplicateStatus.SUSPECTED)
        second.duplicate_status = DuplicateStatus.CONFIRMED_UNIQUE
        self.assertEqual(detector.detect([first, second], 75), [])

    def test_validation_and_reconciliation(self) -> None:
        item = transaction()
        validator = ValidationService()
        status, message = validator.validate(item)
        self.assertEqual(status, ValidationStatus.WARNING)
        self.assertIn("unmatched", message.lower())
        result = validator.reconcile([item], Decimal("10000.00"), Decimal("8749.50"))
        self.assertEqual(result.expected_closing_balance, Decimal("8749.50"))
        self.assertEqual(result.difference, Decimal("0.00"))

    def test_excel_export_contains_transactions_and_summary(self) -> None:
        with TemporaryDirectory() as directory:
            output = Path(directory) / "review.xlsx"
            ExcelExporter().export(output, [transaction()], "sample.pdf")
            workbook = load_workbook(output, data_only=True)
            self.assertEqual(workbook.sheetnames, ["Transactions", "Summary"])
            self.assertEqual(workbook["Transactions"]["C2"].value, "UPI DEMO TRADERS")
            self.assertEqual(workbook["Summary"]["B4"].value, 1)
            workbook.close()

    def test_template_save_and_reload(self) -> None:
        with TemporaryDirectory() as directory:
            service = BankTemplateService(Path(directory))
            mapping = ColumnMapping(transaction_date=0, description=1, debit=2, credit=3)
            saved = service.save_mapping("Demo Bank", mapping)
            self.assertEqual(saved.bank_name, "Demo Bank")
            reloaded = BankTemplateService(Path(directory)).get_by_bank("Demo Bank")
            self.assertIsNotNone(reloaded)
            self.assertEqual(reloaded.mapping.debit, 2)  # type: ignore[union-attr]

    def test_requested_indian_bank_templates_are_available(self) -> None:
        template_directory = Path(__file__).resolve().parents[1] / "config" / "bank_templates"
        templates = BankTemplateService(template_directory).list_templates()
        names = {item.bank_name for item in templates}
        expected = {
            "State Bank of India", "South Indian Bank", "Federal Bank", "HDFC Bank",
            "ICICI Bank", "Axis Bank", "Kotak Mahindra Bank", "Canara Bank",
            "Punjab National Bank", "Bank of Baroda", "Union Bank of India", "Yes Bank",
        }
        self.assertTrue(expected.issubset(names))
        hdfc = next(item for item in templates if item.bank_name == "HDFC Bank")
        self.assertIn("ValueDt", hdfc.column_aliases["value_date"])

    def test_kotak_ifsc_detection_does_not_false_match_sbin_reference(self) -> None:
        template_directory = Path(__file__).resolve().parents[1] / "config" / "bank_templates"
        service = BankTemplateService(template_directory)

        detected = service.detect("IFSC KKBK0009290 NEFT SBIN525198721413")

        self.assertIsNotNone(detected)
        self.assertEqual(detected.template_id, "kotak")  # type: ignore[union-attr]

    def test_bank_name_in_header_outranks_beneficiary_bank_names(self) -> None:
        template_directory = Path(__file__).resolve().parents[1] / "config" / "bank_templates"
        service = BankTemplateService(template_directory)

        detected = service.detect(
            "UNION BANK OF INDIA THRISSUR MAIN STATEMENT OF ACCOUNT "
            + ("transaction details " * 200)
            + "Beneficiary Bank KOTAK MAHINDRA BANK"
        )

        self.assertIsNotNone(detected)
        self.assertEqual(detected.template_id, "union_bank_of_india")  # type: ignore[union-attr]


if __name__ == "__main__":
    unittest.main()
