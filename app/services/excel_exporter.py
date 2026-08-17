"""Formatted, atomic Excel export for reviewed transactions."""

from __future__ import annotations

import re
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.transaction import Transaction

ILLEGAL_EXCEL_CHARACTERS = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")


class ExcelExporter:
    """Write cleaned transaction data and a summary to an `.xlsx` workbook."""

    HEADERS = (
        "Transaction Date", "Value Date", "Description", "Reference Number", "Cheque Number",
        "Debit", "Credit", "Balance", "Voucher Type", "Bank Ledger", "Opposite Ledger",
        "Matching Confidence", "Match Method", "Matching Rule ID", "Narration",
        "Validation Status", "Validation Message", "Duplicate Status", "Tally Voucher Number",
        "Import Status", "Import Message", "Imported At", "Ignored",
    )

    def export(
        self, path: Path, transactions: list[Transaction], source_filename: str = ""
    ) -> Path:
        if path.suffix.lower() != ".xlsx":
            path = path.with_suffix(".xlsx")
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Transactions"
        sheet.append(self.HEADERS)
        header_fill = PatternFill("solid", fgColor="1F4E78")
        for cell in sheet[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for item in transactions:
            sheet.append([
                item.transaction_date, item.value_date, self._clean(item.description),
                self._clean(item.reference_number), self._clean(item.cheque_number),
                item.debit_amount, item.credit_amount, item.balance, item.voucher_type.value,
                self._clean(item.bank_ledger), self._clean(item.opposite_ledger), item.matching_confidence,
                item.matching_method, item.matching_rule_id, self._clean(item.narration),
                item.validation_status.value, self._clean(item.validation_message),
                item.duplicate_status.value, item.tally_voucher_number, item.import_status.value,
                self._clean(item.import_message), item.imported_at, "Yes" if item.ignored else "No",
            ])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        widths = [16, 16, 45, 24, 18, 16, 16, 16, 16, 24, 24, 18, 18, 38, 45, 18, 45, 20, 24, 16, 45, 20, 12]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        for row in sheet.iter_rows(min_row=2):
            row[0].number_format = "dd-mm-yyyy"
            row[1].number_format = "dd-mm-yyyy"
            for cell in row[5:8]:
                cell.number_format = '#,##0.00;[Red]-#,##0.00'
            row[11].number_format = '0"%"'
            row[2].alignment = Alignment(wrap_text=True, vertical="top")
            row[14].alignment = Alignment(wrap_text=True, vertical="top")
            if row[15].value == "Invalid":
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor="FCE4D6")
            elif row[17].value in {"Suspected", "Confirmed duplicate"}:
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor="FFF2CC")
            row[21].number_format = "dd-mm-yyyy hh:mm:ss"

        active = [item for item in transactions if not item.ignored]
        summary = workbook.create_sheet("Summary")
        debit = sum((item.debit_amount for item in active), Decimal("0.00"))
        credit = sum((item.credit_amount for item in active), Decimal("0.00"))
        summary_rows = [
            ("Bank Statement Review Export", ""),
            ("Source file", self._clean(source_filename)),
            ("Exported at", datetime.now()),
            ("Transaction count", len(active)),
            ("Total debit", debit),
            ("Total credit", credit),
            ("Ignored rows", len(transactions) - len(active)),
        ]
        for row in summary_rows:
            summary.append(row)
        summary["A1"].font = Font(size=16, bold=True, color="1F4E78")
        summary.column_dimensions["A"].width = 24
        summary.column_dimensions["B"].width = 45
        summary["B5"].number_format = '#,##0.00'
        summary["B6"].number_format = '#,##0.00'
        temporary = path.with_name(f".{path.name}.tmp")
        workbook.save(temporary)
        temporary.replace(path)
        return path

    @staticmethod
    def _clean(value: str) -> str:
        return ILLEGAL_EXCEL_CHARACTERS.sub("", value)
